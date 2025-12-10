from typing import Iterable, List, Dict, Union, Optional
from io import BytesIO
from openpyxl import load_workbook
import re
import argparse
import json
import sys
import hashlib
import os
from pathlib import Path


def _compute_file_hash(file_path: str) -> str:
    """Compute SHA256 hash of a file for cache key generation."""
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def _get_xlsx_cache_path(file_hash: str, cache_dir: str) -> str:
    """Generate cache file path for Excel extraction."""
    return os.path.join(cache_dir, f"xlsx_{file_hash}.json")


def _load_xlsx_from_cache(
    xlsx_path: str,
    cache_dir: str,
    verbose: bool = False
) -> Optional[List[Dict]]:
    """Load cached Excel extraction results if available."""
    if not os.path.exists(xlsx_path):
        return None
    
    file_hash = _compute_file_hash(xlsx_path)
    cache_path = _get_xlsx_cache_path(file_hash, cache_dir)
    
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cached = json.load(f)
            
            if cached.get("_cache_meta", {}).get("file_hash") == file_hash:
                if verbose:
                    print(f"[Excel] Cache hit! Loading {len(cached.get('elements', []))} elements from cache")
                return cached.get("elements", [])
        except (json.JSONDecodeError, IOError):
            pass
    
    return None


def _save_xlsx_to_cache(
    elements: List[Dict],
    xlsx_path: str,
    cache_dir: str,
    verbose: bool = False
) -> None:
    """Save Excel extraction results to cache."""
    os.makedirs(cache_dir, exist_ok=True)
    
    file_hash = _compute_file_hash(xlsx_path)
    cache_path = _get_xlsx_cache_path(file_hash, cache_dir)
    
    cached = {
        "_cache_meta": {
            "file_hash": file_hash,
            "source_file": os.path.basename(xlsx_path),
            "element_count": len(elements)
        },
        "elements": elements
    }
    
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cached, f, indent=2, ensure_ascii=False)
    
    if verbose:
        print(f"[Excel] Cached {len(elements)} elements to: {cache_path}")


def extract_pi_rows_xlsx(
    source: Union[str, bytes, BytesIO],
    sheets: Iterable[str] = None, # type: ignore
    verbose: bool = False,
    use_cache: bool = True,
    cache_dir: str = ".extraction_cache"
) -> List[Dict]:
    """
    Extract PI calibration elements from an Excel file.
    
    Reads columns L (Ask/Look For) and M (Calibrator notes) to extract
    PI element definitions.
    
    Args:
        source: Path to Excel file, bytes, or BytesIO
        sheets: Optional list of sheet names to process (default: all sheets)
        verbose: Print progress information
        use_cache: Whether to use cached results if available
        cache_dir: Directory for storing cache files
        
    Returns:
        List of dictionaries with PI-Element, Ask/Look For, Calibrator notes
    """
    
    # Check cache if source is a file path
    if use_cache and isinstance(source, str) and os.path.isfile(source):
        cached_result = _load_xlsx_from_cache(source, cache_dir, verbose)
        if cached_result is not None:
            return cached_result
    
    pat = re.compile(
        r'^\s*(?P<num>\d+(?:\.\d+)*)\s*([>\-–—])\s*(?P<text>.+?)\s*$'
    )

    # Load workbook
    if verbose:
        print(f"Loading workbook...")
    
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(BytesIO(source), data_only=True, read_only=True)
    elif isinstance(source, BytesIO):
        wb = load_workbook(source, data_only=True, read_only=True)
    elif isinstance(source, str):
        wb = load_workbook(source, data_only=True, read_only=True)
    else:
        raise TypeError("source must be a path, bytes, or BytesIO")

    out: List[Dict] = []
    rows_processed = 0
    rows_skipped = 0

    worksheets = (
        [wb[s] for s in sheets if s in wb.sheetnames]
        if sheets else wb.worksheets
    )
    
    if verbose:
        print(f"Processing {len(worksheets)} worksheet(s)...")

    for ws in worksheets:
        if verbose:
            print(f"  Reading sheet: {ws.title}")
        
        for row in ws.iter_rows():
            rows_processed += 1

            # Ensure row has at least columns A–M (index 12)
            if len(row) < 13:
                rows_skipped += 1
                continue

            col_L = row[11].value  # column L (12th column)
            col_M = row[12].value  # column M (13th column)

            # ---------------------------
            # UPDATED GUARD
            # Skip row if either column is empty, None, or only whitespace.
            # ---------------------------
            if not col_L or str(col_L).strip() == "":
                rows_skipped += 1
                continue
            if not col_M or str(col_M).strip() == "":
                rows_skipped += 1
                continue

            # Parse L with regex
            match = pat.match(str(col_L))
            if not match:
                rows_skipped += 1
                continue

            num_raw = match.group("num")
            ask_text = match.group("text").strip().strip('"').strip("'")

            # Convert PI-Element to float when possible (e.g., "9.4")
            if num_raw.count(".") <= 1:
                try:
                    pi_value = float(num_raw)
                except ValueError:
                    pi_value = num_raw
            else:
                pi_value = num_raw  # hierarchical like "9.4.1" stays string

            out.append({
                "PI-Element": pi_value,
                "Ask/Look For": ask_text,
                "Calibrator notes": str(col_M).strip()
            })
    
    if verbose:
        print(f"Processed {rows_processed} rows, extracted {len(out)} elements, skipped {rows_skipped} rows")

    wb.close()
    
    # Save to cache if source is a file path
    if use_cache and isinstance(source, str) and os.path.isfile(source):
        _save_xlsx_to_cache(out, source, cache_dir, verbose)
    
    return out


def main():
    """CLI entry point for Excel element extraction."""
    parser = argparse.ArgumentParser(
        description="Extract PI calibration elements from an Excel file (columns L and M)"
    )
    parser.add_argument(
        "input_xlsx",
        help="Path to the input Excel file"
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output JSON file path (default: prints to stdout)"
    )
    parser.add_argument(
        "-s", "--sheets",
        nargs="+",
        default=None,
        help="Specific sheet names to process (default: all sheets)"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Print progress information"
    )
    
    args = parser.parse_args()
    
    # Validate input file exists
    input_path = Path(args.input_xlsx)
    if not input_path.exists():
        print(f"Error: Input file not found: {args.input_xlsx}", file=sys.stderr)
        sys.exit(1)
    
    if not input_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        print(f"Warning: File may not be an Excel file: {input_path.suffix}", file=sys.stderr)
    
    try:
        # Extract elements
        if args.verbose:
            print(f"Input: {args.input_xlsx}")
        
        elements = extract_pi_rows_xlsx(
            source=str(input_path),
            sheets=args.sheets,
            verbose=args.verbose
        )
        
        if not elements:
            print("Warning: No elements extracted. Check that columns L and M contain valid data.", file=sys.stderr)
        
        # Output results
        if args.output:
            output_path = Path(args.output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(elements, f, indent=2, ensure_ascii=False)
            
            if args.verbose:
                print(f"Output: {args.output}")
            print(f"Extracted {len(elements)} PI elements to {args.output}")
        else:
            # Print to stdout
            print(json.dumps(elements, indent=2, ensure_ascii=False))
            
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
